"""openpyxl style constants. All colors derive from docs/design.md tokens."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- Canonical tokens (docs/design.md §Palette) --------------------------

PRIMARY = "0A1F3D"  # --color-primary
SURFACE = "FFFFFF"  # --color-surface
SURFACE_ALT = "F8FAFC"  # --color-surface-alt (zebra)
BORDER = "E2E8F0"  # --color-border
ACCENT = "1E40AF"  # --color-accent
POSITIVE = "15803D"  # --color-positive
NEGATIVE = "B91C1C"  # --color-negative
AMBER = "B45309"  # --color-amber
TEXT = "0F172A"  # --color-text
TEXT_SECONDARY = "64748B"  # --color-text-secondary


def _tint(hex_color: str, alpha: float) -> str:
    """Blend a canonical token against white (cell background) at the given alpha.

    Used to produce conditional-formatting tints from --color-positive /
    --color-negative / --color-amber per docs/design.md §Conditional-formatting
    tints. Returns a 6-digit hex string (no #).
    """
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    nr = round(r * alpha + 255 * (1 - alpha))
    ng = round(g * alpha + 255 * (1 - alpha))
    nb = round(b * alpha + 255 * (1 - alpha))
    return f"{nr:02X}{ng:02X}{nb:02X}"


# --- Color-coding (analyst convention, design.md §Excel export) ----------

INPUT_BLUE = ACCENT  # --color-accent for analyst-editable inputs
HARDCODED_GREEN = POSITIVE  # --color-positive for hardcoded constants

# --- Fill tints (docs/design.md §Conditional-formatting tints) -----------

HEADER_FILL_HEX = TEXT  # --color-text (dark slate) for header band
HEADER_FONT_HEX = SURFACE  # white text on header band
SECTION_HEADER_FILL_HEX = SURFACE_ALT  # zebra row tone for section headers
ANCHOR_TINT_HEX = _tint(PRIMARY, 0.06)  # --color-primary /6 per HANDOFF
QUARTILE_TOP_HEX = _tint(POSITIVE, 0.10)  # /10 per design.md
QUARTILE_BOTTOM_HEX = _tint(NEGATIVE, 0.10)  # /10 per design.md
THRESHOLD_AMBER_HEX = _tint(AMBER, 0.15)  # /15 per design.md
THRESHOLD_RED_HEX = _tint(NEGATIVE, 0.20)  # /20 per design.md

# --- Number formats (docs/design.md §Excel export) -----------------------

NUMFMT_PERCENT = "0.00%;(0.00%)"
NUMFMT_CURRENCY = "$#,##0;($#,##0)"
NUMFMT_INTEGER = "#,##0;(#,##0)"
NUMFMT_DELTA_BPS = '+#,##0 "bps";(#,##0 "bps");"0 bps"'  # numeric Δ for sort/filter

# --- Fonts ---------------------------------------------------------------

FONT_BODY = Font(name="Calibri", size=11)
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_HEX)
FONT_SECTION_HEADER = Font(name="Calibri", size=11, bold=True)
FONT_TITLE = Font(name="Calibri", size=24, bold=True)
FONT_RATIO_NAME = Font(name="Calibri", size=14, bold=True)
FONT_FORMULA_ITALIC = Font(name="Calibri", size=10, italic=True, color=TEXT_SECONDARY)

# --- Fills ---------------------------------------------------------------

FILL_HEADER = PatternFill(fill_type="solid", fgColor=HEADER_FILL_HEX)
FILL_SECTION_HEADER = PatternFill(fill_type="solid", fgColor=SECTION_HEADER_FILL_HEX)
FILL_ANCHOR = PatternFill(fill_type="solid", fgColor=ANCHOR_TINT_HEX)
FILL_QUARTILE_TOP = PatternFill(fill_type="solid", fgColor=QUARTILE_TOP_HEX)
FILL_QUARTILE_BOTTOM = PatternFill(fill_type="solid", fgColor=QUARTILE_BOTTOM_HEX)
FILL_THRESHOLD_AMBER = PatternFill(fill_type="solid", fgColor=THRESHOLD_AMBER_HEX)
FILL_THRESHOLD_RED = PatternFill(fill_type="solid", fgColor=THRESHOLD_RED_HEX)

# --- Borders -------------------------------------------------------------

THIN_GRAY = Side(border_style="thin", color=BORDER)
BORDER_CELL = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

# --- Alignments ----------------------------------------------------------

ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
