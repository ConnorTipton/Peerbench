"""openpyxl writer. Consumes a WorkbookBundle and writes a .xlsx; owns ALL openpyxl wiring."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from peerbench.export import style as st
from peerbench.export.data.types import (
    CompSheetTab,
    CoverTab,
    MethodologyTab,
    RestatementTab,
    SummaryRow,
    SummaryTab,
    TimeSeriesTab,
    WorkbookBundle,
)
from peerbench.export.format import (
    EM_DASH,
    format_delta_bps,
    format_fact_value,
    format_ratio_for_cell,
)
from peerbench.export.quartile import (
    QuartileCutoffs,
    bucket_for_cell,
    compute_quartile_cutoffs,
)


def write_workbook(bundle: WorkbookBundle, out_path: Path) -> Path:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    _write_cover(wb, bundle.cover)
    _write_summary(wb, bundle.summary)
    for sheet in bundle.comp_sheets:
        _write_comp_sheet(wb, sheet)
    for ts in bundle.time_series:
        _write_time_series(wb, ts)
    if bundle.restatement_log is not None:
        _write_restatement(wb, bundle.restatement_log)
    if bundle.methodology is not None:
        _write_methodology(wb, bundle.methodology)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# --- Cover ---------------------------------------------------------------


def _write_cover(wb: Workbook, cover: CoverTab) -> None:
    ws = wb.create_sheet("Cover")
    ws["A1"] = "Peerbench — Bank Peer Benchmarking"
    ws["A1"].font = st.FONT_TITLE
    ws["A2"] = f"{cover.anchor_name} · Cert {cover.anchor_cert}"
    ws["A2"].font = st.FONT_BODY_BOLD
    ws["A3"] = f"As of {cover.quarter_end}"
    ws["A5"] = (
        f"Generated {cover.generated_at.strftime('%Y-%m-%d %H:%M UTC')} "
        f"from data ingested through {cover.data_vintage}"
    )
    ws["A7"] = "Workbook contents:"
    ws["A7"].font = st.FONT_BODY_BOLD
    contents = [
        ("Summary", "all 30 ratios, anchor + peers, latest quarter"),
        ("Comp Sheets", "one tab per peer, side-by-side I/S + B/S + ratios"),
        ("Time Series", "8 quarters by ratio category"),
        ("Restatement Log", "facts revised by FDIC affecting workbook ratios"),
        ("Methodology", "formulas, sources, regulatory thresholds"),
    ]
    for i, (label, desc) in enumerate(contents):
        ws.cell(row=8 + i, column=1, value=f"• {label} — {desc}")
    ws["A14"] = "Data sources: FDIC BankFind API, FFIEC CDR bulk files."
    ws["A15"] = "Restatements detected automatically; see Restatement Log tab."
    next_row = 17
    for note in cover.notes:
        cell = ws.cell(row=next_row, column=1, value=note)
        cell.font = st.FONT_BODY_BOLD
        next_row += 1
    ws.column_dimensions["A"].width = 100


# --- Summary -------------------------------------------------------------


def _summary_cell_fill(
    *,
    value: object,
    is_anchor: bool,
    amber_pct: object,
    red_pct: object,
    cutoffs: QuartileCutoffs | None,
    direction: str,
) -> PatternFill | None:
    """Layer precedence: regulatory red > amber > quartile > anchor."""
    if value is not None:
        if red_pct is not None and value >= red_pct:  # type: ignore[operator]
            return st.FILL_THRESHOLD_RED
        if amber_pct is not None and value >= amber_pct:  # type: ignore[operator]
            return st.FILL_THRESHOLD_AMBER
        if not is_anchor:
            bucket = bucket_for_cell(value, cutoffs, direction)  # type: ignore[arg-type]
            if bucket == "top":
                return st.FILL_QUARTILE_TOP
            if bucket == "bottom":
                return st.FILL_QUARTILE_BOTTOM
    if is_anchor:
        return st.FILL_ANCHOR
    return None


def _write_summary(wb: Workbook, summary: SummaryTab) -> None:
    ws = wb.create_sheet("Summary")
    cols = summary.institution_columns
    n_inst = len(cols)
    median_col = 2 + n_inst + 1
    rank_col = median_col + 1
    delta_col = rank_col + 1

    header_row1 = ["Category", "Ratio"]
    header_row2 = ["", ""]
    for cert, name in cols:
        header_row1.append(name)
        header_row2.append(f"Cert {cert}")
    header_row1.extend(["Peer median", "Anchor rank", "Δ vs median"])
    header_row2.extend(["", "", ""])

    for col_idx, val in enumerate(header_row1, start=1):
        cell = ws.cell(row=1, column=col_idx, value=val)
        cell.font = st.FONT_HEADER
        cell.fill = st.FILL_HEADER
        cell.alignment = st.ALIGN_CENTER
    for col_idx, val in enumerate(header_row2, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = st.FONT_BODY
        cell.alignment = st.ALIGN_CENTER

    current_row = 3
    last_category: str | None = None
    for row_data in summary.rows:
        if row_data.category != last_category:
            cat_cell = ws.cell(row=current_row, column=1, value=row_data.category)
            cat_cell.font = st.FONT_SECTION_HEADER
            cat_cell.fill = st.FILL_SECTION_HEADER
            ws.merge_cells(
                start_row=current_row,
                end_row=current_row,
                start_column=1,
                end_column=delta_col,
            )
            current_row += 1
            last_category = row_data.category
        _write_summary_data_row(
            ws=ws,
            row=current_row,
            row_data=row_data,
            cols=cols,
            median_col=median_col,
            rank_col=rank_col,
            delta_col=delta_col,
        )
        current_row += 1

    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    for i in range(3, 3 + n_inst + 3):
        ws.column_dimensions[get_column_letter(i)].width = 14


def _write_summary_data_row(
    *,
    ws: object,
    row: int,
    row_data: SummaryRow,
    cols: list[tuple[int, str]],
    median_col: int,
    rank_col: int,
    delta_col: int,
) -> None:
    ws.cell(row=row, column=1, value="")  # type: ignore[attr-defined]
    ws.cell(row=row, column=2, value=row_data.display_name).font = st.FONT_BODY  # type: ignore[attr-defined]

    peer_numeric = [row_data.peer_values.get(cert) for cert, _ in cols[1:]]
    cutoffs = compute_quartile_cutoffs(peer_numeric)

    for inst_idx, (cert, _) in enumerate(cols):
        col_idx = 3 + inst_idx
        value = row_data.anchor_value if inst_idx == 0 else row_data.peer_values.get(cert)
        cell = ws.cell(  # type: ignore[attr-defined]
            row=row,
            column=col_idx,
            value=format_ratio_for_cell(value),
        )
        cell.number_format = st.NUMFMT_PERCENT
        cell.alignment = st.ALIGN_RIGHT
        fill = _summary_cell_fill(
            value=value,
            is_anchor=inst_idx == 0,
            amber_pct=row_data.amber_pct,
            red_pct=row_data.red_pct,
            cutoffs=cutoffs,
            direction=row_data.direction,
        )
        if fill is not None:
            cell.fill = fill

    median_cell = ws.cell(  # type: ignore[attr-defined]
        row=row,
        column=median_col,
        value=format_ratio_for_cell(row_data.peer_median),
    )
    median_cell.number_format = st.NUMFMT_PERCENT
    median_cell.alignment = st.ALIGN_RIGHT

    rank_cell = ws.cell(row=row, column=rank_col, value=row_data.anchor_rank)  # type: ignore[attr-defined]
    rank_cell.number_format = st.NUMFMT_INTEGER
    rank_cell.alignment = st.ALIGN_RIGHT

    ws.cell(  # type: ignore[attr-defined]
        row=row,
        column=delta_col,
        value=format_delta_bps(row_data.anchor_value, row_data.peer_median),
    ).alignment = st.ALIGN_RIGHT


# --- Comp Sheet ----------------------------------------------------------


def _write_comp_sheet(wb: Workbook, sheet: CompSheetTab) -> None:
    ws = wb.create_sheet(sheet.sheet_name)
    ws["A1"] = f"{sheet.anchor_name} vs {sheet.peer_name} · {sheet.quarter_id}"  # type: ignore[index]
    ws["A1"].font = st.FONT_TITLE  # type: ignore[index]

    # Section A: Income statement
    row = 3
    ws.cell(
        row=row, column=1, value="Income statement (YTD, $ thousands)"
    ).font = st.FONT_RATIO_NAME  # type: ignore[attr-defined]
    row += 1
    n_q = len(sheet.income_statement_quarters)
    ws.cell(row=row, column=2, value=sheet.anchor_name).font = st.FONT_HEADER  # type: ignore[attr-defined]
    ws.cell(row=row, column=2 + n_q, value=sheet.peer_name).font = st.FONT_HEADER  # type: ignore[attr-defined]
    ws.cell(row=row, column=2 + 2 * n_q, value="Δ (current)").font = st.FONT_HEADER  # type: ignore[attr-defined]
    row += 1
    for q_idx, q in enumerate(sheet.income_statement_quarters):
        ws.cell(row=row, column=2 + q_idx, value=q).font = st.FONT_BODY_BOLD  # type: ignore[attr-defined]
        ws.cell(row=row, column=2 + n_q + q_idx, value=q).font = st.FONT_BODY_BOLD  # type: ignore[attr-defined]
    row += 1
    for line in sheet.income_statement:
        ws.cell(row=row, column=1, value=line.label).font = st.FONT_BODY  # type: ignore[attr-defined]
        for q_idx, v in enumerate(line.anchor_values):
            c = ws.cell(row=row, column=2 + q_idx, value=float(v) if v is not None else None)  # type: ignore[attr-defined]
            c.number_format = st.NUMFMT_CURRENCY
            c.alignment = st.ALIGN_RIGHT
        for q_idx, v in enumerate(line.peer_values):
            c = ws.cell(row=row, column=2 + n_q + q_idx, value=float(v) if v is not None else None)  # type: ignore[attr-defined]
            c.number_format = st.NUMFMT_CURRENCY
            c.alignment = st.ALIGN_RIGHT
        a_curr = line.anchor_values[-1] if line.anchor_values else None
        p_curr = line.peer_values[-1] if line.peer_values else None
        delta = float(a_curr - p_curr) if a_curr is not None and p_curr is not None else None
        c = ws.cell(row=row, column=2 + 2 * n_q, value=delta)  # type: ignore[attr-defined]
        c.number_format = st.NUMFMT_CURRENCY
        c.alignment = st.ALIGN_RIGHT
        row += 1

    # Section B: Balance sheet
    row += 1
    ws.cell(
        row=row, column=1, value="Balance sheet (period-end, $ thousands)"
    ).font = st.FONT_RATIO_NAME  # type: ignore[attr-defined]
    row += 1
    ws.cell(row=row, column=2, value=sheet.anchor_name).font = st.FONT_HEADER  # type: ignore[attr-defined]
    ws.cell(row=row, column=3, value=sheet.peer_name).font = st.FONT_HEADER  # type: ignore[attr-defined]
    ws.cell(row=row, column=4, value="Δ").font = st.FONT_HEADER  # type: ignore[attr-defined]
    row += 1
    for bs_line in sheet.balance_sheet:
        ws.cell(row=row, column=1, value=bs_line.label)  # type: ignore[attr-defined]
        c = ws.cell(
            row=row,
            column=2,
            value=float(bs_line.anchor_value) if bs_line.anchor_value is not None else None,
        )  # type: ignore[attr-defined]
        c.number_format = st.NUMFMT_CURRENCY
        c.alignment = st.ALIGN_RIGHT
        c = ws.cell(
            row=row,
            column=3,
            value=float(bs_line.peer_value) if bs_line.peer_value is not None else None,
        )  # type: ignore[attr-defined]
        c.number_format = st.NUMFMT_CURRENCY
        c.alignment = st.ALIGN_RIGHT
        delta = (
            float(bs_line.anchor_value - bs_line.peer_value)
            if bs_line.anchor_value is not None and bs_line.peer_value is not None
            else None
        )
        c = ws.cell(row=row, column=4, value=delta)  # type: ignore[attr-defined]
        c.number_format = st.NUMFMT_CURRENCY
        c.alignment = st.ALIGN_RIGHT
        row += 1

    # Section C: Ratios block
    row += 1
    ws.cell(row=row, column=1, value="Ratios").font = st.FONT_RATIO_NAME  # type: ignore[attr-defined]
    row += 1
    headers = ["Ratio", "Formula", sheet.anchor_name, sheet.peer_name, "Δ", "Direction"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)  # type: ignore[attr-defined]
        cell.font = st.FONT_HEADER
        cell.fill = st.FILL_HEADER
    row += 1
    last_cat: str | None = None
    for r in sheet.ratios:
        if r.category != last_cat:
            cat_cell = ws.cell(row=row, column=1, value=r.category)  # type: ignore[attr-defined]
            cat_cell.font = st.FONT_SECTION_HEADER
            cat_cell.fill = st.FILL_SECTION_HEADER
            ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=6)  # type: ignore[attr-defined]
            row += 1
            last_cat = r.category
        ws.cell(row=row, column=1, value=r.display_name)  # type: ignore[attr-defined]
        ws.cell(row=row, column=2, value=r.formula).font = st.FONT_FORMULA_ITALIC  # type: ignore[attr-defined]
        a_cell = ws.cell(row=row, column=3, value=format_ratio_for_cell(r.anchor_value))  # type: ignore[attr-defined]
        a_cell.number_format = st.NUMFMT_PERCENT
        a_cell.alignment = st.ALIGN_RIGHT
        p_cell = ws.cell(row=row, column=4, value=format_ratio_for_cell(r.peer_value))  # type: ignore[attr-defined]
        p_cell.number_format = st.NUMFMT_PERCENT
        p_cell.alignment = st.ALIGN_RIGHT
        ws.cell(
            row=row, column=5, value=format_delta_bps(r.anchor_value, r.peer_value)
        ).alignment = st.ALIGN_RIGHT  # type: ignore[attr-defined]
        direction_text = {
            "higher_is_positive": "Higher better",
            "higher_is_negative": "Lower better",
            "neutral": "Neutral",
        }[r.direction]
        ws.cell(row=row, column=6, value=direction_text)  # type: ignore[attr-defined]
        row += 1

    ws.freeze_panes = "B4"  # type: ignore[attr-defined]
    ws.column_dimensions["A"].width = 28  # type: ignore[attr-defined]
    for c_idx in range(2, 12):
        ws.column_dimensions[get_column_letter(c_idx)].width = 14  # type: ignore[attr-defined]


# --- Time Series ---------------------------------------------------------


def _write_time_series(wb: Workbook, tab: TimeSeriesTab) -> None:
    ws = wb.create_sheet(tab.sheet_name)
    row = 1
    for block in tab.blocks:
        ws.cell(row=row, column=1, value=block.display_name).font = st.FONT_RATIO_NAME  # type: ignore[attr-defined]
        row += 1
        ws.cell(row=row, column=1, value=block.formula).font = st.FONT_FORMULA_ITALIC  # type: ignore[attr-defined]
        row += 1
        for q_idx, q in enumerate(block.quarter_ids):
            cell = ws.cell(row=row, column=2 + q_idx, value=q)  # type: ignore[attr-defined]
            cell.font = st.FONT_HEADER
            cell.fill = st.FILL_HEADER
        row += 1
        cutoffs_by_q = [
            compute_quartile_cutoffs([r[2][q_idx] for r in block.rows])
            for q_idx in range(len(block.quarter_ids))
        ]
        for inst_idx, (_cert, name, values) in enumerate(block.rows):
            label_cell = ws.cell(row=row, column=1, value=name)  # type: ignore[attr-defined]
            if inst_idx == 0:
                label_cell.fill = st.FILL_ANCHOR
                label_cell.font = st.FONT_BODY_BOLD
            for q_idx, v in enumerate(values):
                cell = ws.cell(  # type: ignore[attr-defined]
                    row=row,
                    column=2 + q_idx,
                    value=format_ratio_for_cell(v),
                )
                cell.number_format = st.NUMFMT_PERCENT
                cell.alignment = st.ALIGN_RIGHT
                if inst_idx == 0:
                    cell.fill = st.FILL_ANCHOR
                else:
                    bucket = bucket_for_cell(v, cutoffs_by_q[q_idx], block.direction)
                    if bucket == "top":
                        cell.fill = st.FILL_QUARTILE_TOP
                    elif bucket == "bottom":
                        cell.fill = st.FILL_QUARTILE_BOTTOM
            row += 1
        row += 1  # blank spacer
    ws.freeze_panes = "B1"  # type: ignore[attr-defined]
    ws.column_dimensions["A"].width = 28  # type: ignore[attr-defined]
    for c_idx in range(2, 12):
        ws.column_dimensions[get_column_letter(c_idx)].width = 12  # type: ignore[attr-defined]


# --- Restatement Log -----------------------------------------------------


def _write_restatement(wb: Workbook, tab: RestatementTab) -> None:
    ws = wb.create_sheet("Restatement Log")
    headers = [
        "Detected at",
        "Cert",
        "Bank",
        "Quarter",
        "Field",
        "Old value",
        "New value",
        "Δ",
        "Affected ratios",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)  # type: ignore[attr-defined]
        cell.font = st.FONT_HEADER
        cell.fill = st.FILL_HEADER
    if not tab.rows:
        ws.cell(  # type: ignore[attr-defined]
            row=2,
            column=1,
            value="No restatements affecting workbook ratios.",
        ).font = st.FONT_FORMULA_ITALIC
    else:
        for row_idx, r in enumerate(tab.rows, start=2):
            ws.cell(row=row_idx, column=1, value=r.detected_at.strftime("%Y-%m-%d"))  # type: ignore[attr-defined]
            ws.cell(row=row_idx, column=2, value=r.cert)  # type: ignore[attr-defined]
            ws.cell(row=row_idx, column=3, value=r.bank_name)  # type: ignore[attr-defined]
            ws.cell(row=row_idx, column=4, value=r.quarter_id)  # type: ignore[attr-defined]
            ws.cell(row=row_idx, column=5, value=r.field_code)  # type: ignore[attr-defined]
            ws.cell(
                row=row_idx, column=6, value=format_fact_value(r.old_value)
            ).alignment = st.ALIGN_RIGHT  # type: ignore[attr-defined]
            ws.cell(
                row=row_idx, column=7, value=format_fact_value(r.new_value)
            ).alignment = st.ALIGN_RIGHT  # type: ignore[attr-defined]
            delta = (
                r.new_value - r.old_value
                if r.old_value is not None and r.new_value is not None
                else None
            )
            ws.cell(
                row=row_idx, column=8, value=format_fact_value(delta)
            ).alignment = st.ALIGN_RIGHT  # type: ignore[attr-defined]
            ws.cell(row=row_idx, column=9, value=", ".join(r.affected_ratios))  # type: ignore[attr-defined]
    ws.freeze_panes = "A2"  # type: ignore[attr-defined]
    widths = [14, 8, 28, 10, 14, 14, 14, 12, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w  # type: ignore[attr-defined]


# --- Methodology ---------------------------------------------------------


def _write_methodology(wb: Workbook, tab: MethodologyTab) -> None:
    ws = wb.create_sheet("Methodology")
    ws["A1"] = "Methodology"  # type: ignore[index]
    ws["A1"].font = st.FONT_TITLE  # type: ignore[index]
    row = 3
    for note in tab.intro_notes:
        ws.cell(row=row, column=1, value=note.label).font = st.FONT_BODY_BOLD  # type: ignore[attr-defined]
        ws.cell(row=row, column=2, value=note.text).alignment = Alignment(wrap_text=True)  # type: ignore[attr-defined]
        row += 1
    row += 1
    for block in tab.blocks:
        ws.cell(row=row, column=1, value=block.display_name).font = st.FONT_RATIO_NAME  # type: ignore[attr-defined]
        ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=4)  # type: ignore[attr-defined]
        row += 1
        keys: list[tuple[str, str | None]] = [
            ("Category", block.category),
            ("Formula", block.formula),
            ("Source fields", ", ".join(block.source_fields) or EM_DASH),
            ("Annualization", block.annualization),
            ("Basis", block.basis),
            ("FDIC pre-computed", block.fdic_precomputed or EM_DASH),
            ("Regulatory threshold", block.regulatory_threshold or EM_DASH),
            ("Notes", block.notes or EM_DASH),
        ]
        for label, value in keys:
            ws.cell(row=row, column=1, value=label).font = st.FONT_BODY_BOLD  # type: ignore[attr-defined]
            ws.cell(row=row, column=2, value=value).alignment = Alignment(wrap_text=True)  # type: ignore[attr-defined]
            row += 1
        row += 1
    ws.freeze_panes = "A2"  # type: ignore[attr-defined]
    ws.column_dimensions["A"].width = 22  # type: ignore[attr-defined]
    ws.column_dimensions["B"].width = 80  # type: ignore[attr-defined]
