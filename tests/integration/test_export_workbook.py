"""Round-trip integration: build a real .xlsx and read it back."""

from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import JSON, create_engine
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.orm import Session

from peerbench.db import (
    Base,
    Fact,
    Institution,
    QualityLog,
    Quarter,
    Ratio,
    RatioDef,
)
from peerbench.export import run_export


def _patch_jsonb_for_sqlite() -> None:
    """Swap JSONB columns to JSON so SQLite can create the schema.

    Base.metadata is process-global; the swap is idempotent (JSON → JSON is a no-op)
    and harmless because these tests run against an in-memory SQLite engine that is
    discarded after each test.  The production code path uses Postgres, which handles
    JSONB natively.
    """
    for table in Base.metadata.tables.values():
        for col in table.columns:
            if isinstance(col.type, JSONB):
                col.type = JSON()


@pytest.mark.integration
def test_export_writes_six_tab_kinds(tmp_path: Path) -> None:
    _patch_jsonb_for_sqlite()
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        _seed(session)
        out_path = run_export(
            session,
            anchor_cert=4063,
            quarter_id="2025-Q4",
            out_dir=tmp_path,
        )

    assert out_path.exists()
    wb = load_workbook(out_path)
    names = wb.sheetnames
    assert "Cover" in names
    assert "Summary" in names
    assert "Restatement Log" in names
    assert "Methodology" in names
    assert any(n in {"Peer A", "Peer B"} for n in names)
    assert "Profitability" in names

    summary = wb["Summary"]
    assert summary.freeze_panes == "C3"
    for row in summary.iter_rows(min_row=3, max_row=15, values_only=False):
        if row[1].value == "Net Interest Margin":
            assert row[2].value == pytest.approx(0.0342, abs=1e-6)
            assert row[2].number_format == "0.00%;(0.00%)"
            break
    else:
        pytest.fail("Net Interest Margin row not found on Summary")

    rl = wb["Restatement Log"]
    assert rl["E2"].value == "NETINC"

    meth = wb["Methodology"]
    found = False
    for row in meth.iter_rows(values_only=True):
        if any(v == "Net Interest Margin" for v in row if v is not None):
            found = True
            break
    assert found, "Methodology missing NIM block"


def _seed(session: Session) -> None:
    session.add(Institution(cert=4063, name="MidFirst", active=True))
    session.add(Institution(cert=8001, name="Peer A", active=True))
    session.add(Institution(cert=8002, name="Peer B", active=True))
    session.add(Institution(cert=8003, name="Peer C", active=True))
    session.add(Institution(cert=8004, name="Peer D", active=True))

    quarter_ids = [
        "2024-Q1",
        "2024-Q2",
        "2024-Q3",
        "2024-Q4",
        "2025-Q1",
        "2025-Q2",
        "2025-Q3",
        "2025-Q4",
    ]
    for qid in quarter_ids:
        year, q = int(qid[:4]), int(qid[-1])
        report_date = _quarter_end(year, q)
        session.add(
            Quarter(
                quarter_id=qid,
                year=year,
                quarter=q,
                report_date=report_date,
                ingest_at=datetime(2026, 5, 22, tzinfo=UTC),
                source="fdic_api",
            )
        )

    session.add(
        RatioDef(
            ratio_id="nim",
            display_name="Net Interest Margin",
            category="profitability",
            numerator_formula="Net interest income",
            denominator_formula="Avg earning assets",
            annualize=True,
            avg_or_eop="AVG",
            fdic_precomputed_code="NIMY",
            ubpr_concept=None,
            regulatory_threshold=None,
            suppress_when=None,
            notes=None,
        )
    )

    anchor_vals = {"2025-Q4": Decimal("0.0342")}
    peer_offsets = {
        8001: Decimal("-0.002"),
        8002: Decimal("-0.004"),
        8003: Decimal("-0.006"),
        8004: Decimal("-0.008"),
    }
    for qid in quarter_ids:
        anchor_val = anchor_vals.get(qid, Decimal("0.0320"))
        session.add(
            Ratio(
                cert=4063,
                quarter_id=qid,
                ratio_id="nim",
                value=anchor_val,
                formula_version="v1",
                data_quality="ok",
                computed_at=datetime(2026, 5, 22, tzinfo=UTC),
            )
        )
        for cert, offset in peer_offsets.items():
            session.add(
                Ratio(
                    cert=cert,
                    quarter_id=qid,
                    ratio_id="nim",
                    value=anchor_val + offset,
                    formula_version="v1",
                    data_quality="ok",
                    computed_at=datetime(2026, 5, 22, tzinfo=UTC),
                )
            )

    income_codes = [
        "INTINC",
        "EINTEXP",
        "NIM",
        "ELNATR",
        "NONII",
        "NONIX",
        "NETINC",
    ]
    balance_codes = ["ASSET", "LNLSGR", "SC", "CHBAL", "DEP", "LIAB", "EQ"]
    for cert in [4063, 8001, 8002, 8003, 8004]:
        for code in income_codes + balance_codes:
            for qid in quarter_ids:
                session.add(
                    Fact(
                        cert=cert,
                        quarter_id=qid,
                        field_code=code,
                        value=Decimal("1000000"),
                        first_seen_at=datetime(2026, 5, 22, tzinfo=UTC),
                        last_updated_at=datetime(2026, 5, 22, tzinfo=UTC),
                    )
                )

    session.add(
        QualityLog(
            cert=4063,
            quarter_id="2025-Q3",
            field_code="NETINC",
            event_type="restated",
            old_value=Decimal("100"),
            new_value=Decimal("120"),
            detected_at=datetime(2026, 5, 22, tzinfo=UTC),
        )
    )
    session.commit()


def _quarter_end(year: int, q: int) -> date:
    # Q1 → Mar 31; Q2 → Jun 30; Q3 → Sep 30; Q4 → Dec 31.
    return {
        1: date(year, 3, 31),
        2: date(year, 6, 30),
        3: date(year, 9, 30),
        4: date(year, 12, 31),
    }[q]
